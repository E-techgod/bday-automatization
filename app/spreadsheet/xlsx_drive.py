from __future__ import annotations

from collections.abc import Callable
from io import BytesIO
from typing import Any
from zipfile import BadZipFile

from googleapiclient.errors import HttpError  # type: ignore[import-untyped]
from openpyxl import load_workbook  # type: ignore[import-untyped]
from openpyxl.utils.exceptions import (  # type: ignore[import-untyped]
    InvalidFileException,
)

from app.config import Config
from app.retry import retry_with_backoff
from app.spreadsheet.base import (
    SpreadsheetError,
    SpreadsheetProvider,
    build_row_dict,
    resolve_headers,
)

DriveServiceFactory = Callable[[], Any]
_TRANSIENT_HTTP_STATUSES = frozenset({429, 500, 502, 503, 504})


class _RetryableSpreadsheetError(Exception):
    """Wrap transient API failures so the retry decorator can target them."""


class XlsxDriveProvider(SpreadsheetProvider):
    def __init__(self, config: Config, service_factory: DriveServiceFactory) -> None:
        self._config = config
        self._service_factory = service_factory

    def load_rows(self) -> list[dict[str, object]]:
        workbook_bytes = self._download_workbook_bytes_with_retry()
        try:
            workbook = load_workbook(workbook_bytes, data_only=True, read_only=True)
        except (
            BadZipFile,
            InvalidFileException,
            KeyError,
            ValueError,
        ) as exc:
            raise SpreadsheetError(
                "Downloaded file is not a valid .xlsx workbook"
            ) from exc
        try:
            worksheet = workbook.worksheets[0]
            row_iterator = worksheet.iter_rows(values_only=True)
            try:
                raw_header_row = next(row_iterator)
            except StopIteration:
                return []

            header_row = ["" if cell is None else str(cell) for cell in raw_header_row]
            resolved_headers = resolve_headers(header_row, self._config)
            return [
                build_row_dict(list(raw_row), resolved_headers)
                for raw_row in row_iterator
            ]
        finally:
            workbook.close()

    def _download_workbook_bytes_with_retry(self) -> BytesIO:
        @retry_with_backoff(
            max_attempts=self._config.retry_max_attempts,
            base_delay_seconds=self._config.retry_base_delay_seconds,
            retryable_exceptions=(
                _RetryableSpreadsheetError,
                TimeoutError,
                ConnectionError,
            ),
        )
        def download_workbook_bytes() -> BytesIO:
            return self._download_workbook_bytes()

        return download_workbook_bytes()

    def _download_workbook_bytes(self) -> BytesIO:
        service = self._service_factory()
        try:
            request = service.files().get_media(
                fileId=self._config.google_drive_file_id
            )
            workbook_buffer = BytesIO(request.execute())
        except HttpError as exc:
            if exc.resp.status in _TRANSIENT_HTTP_STATUSES:
                raise _RetryableSpreadsheetError("Transient Drive API failure") from exc
            raise SpreadsheetError("Drive API request failed") from exc
        workbook_buffer.seek(0)
        return workbook_buffer
