from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from unittest.mock import Mock, patch

import pytest
from openpyxl import Workbook

from app.birthday_rules import parse_birthday
from app.config import Config
from app.spreadsheet.base import SpreadsheetError
from app.spreadsheet.base import resolve_headers as base_resolve_headers
from app.spreadsheet.xlsx_drive import XlsxDriveProvider


def test_xlsx_drive_load_rows_maps_headers_with_normalization() -> None:
    provider = XlsxDriveProvider(
        config=_build_config(),
        service_factory=lambda: _FakeDriveService(
            _build_workbook_bytes(
                [" name ", "EMAIL", " birthday ", " last birthday email year "],
                ["Test Person", "test.person@example.com", "2000-01-02", 2025],
            )
        ),
    )

    rows = provider.load_rows()

    assert rows == [
        {
            "Name": "Test Person",
            "Email": "test.person@example.com",
            "Birthday": "2000-01-02",
            "Last Birthday Email Year": 2025,
        }
    ]


def test_xlsx_drive_load_rows_raises_for_missing_required_header() -> None:
    provider = XlsxDriveProvider(
        config=_build_config(),
        service_factory=lambda: _FakeDriveService(
            _build_workbook_bytes(
                ["Name", "Email"],
                ["Test Person", "test.person@example.com"],
            )
        ),
    )

    with pytest.raises(
        SpreadsheetError, match="Missing required spreadsheet header: Birthday"
    ):
        provider.load_rows()


def test_xlsx_drive_load_rows_allows_missing_optional_last_sent_year() -> None:
    provider = XlsxDriveProvider(
        config=_build_config(),
        service_factory=lambda: _FakeDriveService(
            _build_workbook_bytes(
                ["Name", "Email", "Birthday"],
                ["Test Person", "test.person@example.com", "2000-01-02"],
            )
        ),
    )

    rows = provider.load_rows()

    assert rows == [
        {
            "Name": "Test Person",
            "Email": "test.person@example.com",
            "Birthday": "2000-01-02",
        }
    ]


def test_xlsx_drive_load_rows_allows_blank_header_spacer_columns() -> None:
    provider = XlsxDriveProvider(
        config=_build_config(),
        service_factory=lambda: _FakeDriveService(
            _build_workbook_bytes(
                ["Name", "Email", "Birthday", None, None, "Notes"],
                [
                    "Test Person",
                    "test.person@example.com",
                    "2000-01-02",
                    "ignored-1",
                    "ignored-2",
                    "Important note",
                ],
            )
        ),
    )

    rows = provider.load_rows()

    assert rows == [
        {
            "Name": "Test Person",
            "Email": "test.person@example.com",
            "Birthday": "2000-01-02",
        }
    ]


def test_xlsx_drive_load_rows_preserves_native_birthday_datetime() -> None:
    birthday_value = date(2000, 1, 2)
    provider = XlsxDriveProvider(
        config=_build_config(),
        service_factory=lambda: _FakeDriveService(
            _build_workbook_bytes(
                ["Name", "Email", "Birthday"],
                ["  Test Person  ", " test.person@example.com ", birthday_value],
            )
        ),
    )

    rows = provider.load_rows()

    assert rows[0]["Name"] == "Test Person"
    assert rows[0]["Email"] == "test.person@example.com"
    assert isinstance(rows[0]["Birthday"], date | datetime)
    loaded_birthday = rows[0]["Birthday"]
    if isinstance(loaded_birthday, datetime):
        assert loaded_birthday.date() == birthday_value
    else:
        assert loaded_birthday == birthday_value
    assert parse_birthday(rows[0]["Birthday"]) == date(2000, 1, 2)


def test_xlsx_drive_load_rows_closes_workbook_on_header_resolution_failure() -> None:
    workbook = Mock()
    worksheet = Mock()
    worksheet.iter_rows.return_value = iter([("Name", "Email")])
    workbook.worksheets = [worksheet]

    provider = XlsxDriveProvider(
        config=_build_config(),
        service_factory=lambda: _FakeDriveService(b"synthetic-workbook-bytes"),
    )

    with (
        patch("app.spreadsheet.xlsx_drive.load_workbook", return_value=workbook),
        pytest.raises(
            SpreadsheetError, match="Missing required spreadsheet header: Birthday"
        ),
    ):
        provider.load_rows()

    workbook.close.assert_called_once_with()


def test_xlsx_drive_load_rows_raises_spreadsheet_error_for_invalid_workbook() -> None:
    provider = XlsxDriveProvider(
        config=_build_config(),
        service_factory=lambda: _FakeDriveService(b"not a real xlsx file"),
    )

    with pytest.raises(
        SpreadsheetError, match="Downloaded file is not a valid .xlsx workbook"
    ):
        provider.load_rows()


def test_xlsx_drive_load_rows_consumes_only_header_before_resolving() -> None:
    workbook = Mock()
    worksheet = Mock()
    yielded_rows: list[tuple[object, ...]] = []

    def iter_rows(*, values_only: bool) -> object:
        assert values_only is True

        def row_generator() -> object:
            for row in (
                (" Name ", "Email", "Birthday"),
                ("Test Person", "test.person@example.com", "2000-01-02"),
                ("Second Person", "second@example.com", "1999-05-06"),
            ):
                yielded_rows.append(row)
                yield row

        return row_generator()

    worksheet.iter_rows.side_effect = iter_rows
    workbook.worksheets = [worksheet]

    provider = XlsxDriveProvider(
        config=_build_config(),
        service_factory=lambda: _FakeDriveService(b"synthetic-workbook-bytes"),
    )

    def resolving_headers(header_row: list[str], config: Config) -> dict[str, int]:
        assert yielded_rows == [(" Name ", "Email", "Birthday")]
        return base_resolve_headers(header_row, config)

    with (
        patch("app.spreadsheet.xlsx_drive.load_workbook", return_value=workbook),
        patch(
            "app.spreadsheet.xlsx_drive.resolve_headers", side_effect=resolving_headers
        ),
    ):
        rows = provider.load_rows()

    assert rows == [
        {
            "Name": "Test Person",
            "Email": "test.person@example.com",
            "Birthday": "2000-01-02",
        },
        {
            "Name": "Second Person",
            "Email": "second@example.com",
            "Birthday": "1999-05-06",
        },
    ]
    workbook.close.assert_called_once_with()


class _FakeDriveService:
    def __init__(self, workbook_bytes: bytes) -> None:
        self._workbook_bytes = workbook_bytes

    def files(self) -> _FakeDriveService:
        return self

    def get_media(self, fileId: str) -> _FakeDriveRequest:
        assert fileId == "test-drive-id"
        return _FakeDriveRequest(self._workbook_bytes)


class _FakeDriveRequest:
    def __init__(self, workbook_bytes: bytes) -> None:
        self._workbook_bytes = workbook_bytes

    def execute(self) -> bytes:
        return self._workbook_bytes


def _build_workbook_bytes(header_row: list[object], data_row: list[object]) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(header_row)
    worksheet.append(data_row)

    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


def _build_config() -> Config:
    return Config(
        app_timezone="America/Chicago",
        dry_run=True,
        test_date=None,
        spreadsheet_mode="xlsx_drive",
        google_sheet_id="test-sheet-id",
        google_sheet_tab="Birthdays",
        google_drive_file_id="test-drive-id",
        name_column="Name",
        email_column="Email",
        birthday_column="Birthday",
        last_sent_year_column="Last Birthday Email Year",
        email_provider="gmail",
        email_from_name="Test Sender",
        email_from_address="sender@example.com",
        email_subject_template="Happy Birthday, {{name}}!",
        google_credentials_file=Path("synthetic-credentials.json"),
        google_impersonate_subject="sender@example.com",
        birthday_image_mode="none",
        birthday_image_path=Path("synthetic-banner.png"),
        birthday_image_url="",
        birthday_image_alt="Happy Birthday",
        birthday_image_width=600,
        state_db_path=Path("synthetic-state.db"),
        stale_claim_timeout_minutes=30,
        retry_max_attempts=3,
        retry_base_delay_seconds=1.0,
        log_level="INFO",
    )
